"""Unified FastAPI backend for Tawiza Dashboard.

Provides REST API endpoints for the dashboard frontend.
Runs alongside the MCP server on a different port.

Usage:
    python -m src.infrastructure.dashboard.api --port 3001
"""

import json
import os
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel

from .database import DashboardDB, init_database
from .models import Analysis, WatchItem
from .stats import get_async_stats

# Initialize FastAPI app
app = FastAPI(
    title="Tawiza Dashboard API",
    description="REST API for Tawiza Dashboard - Market Analysis Platform",
    version="1.0.0",
)

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database instance
db: DashboardDB | None = None


def get_db() -> DashboardDB:
    """Get database instance, initializing if needed."""
    global db
    if db is None:
        db_path = Path.home() / ".tawiza" / "dashboard.db"
        db_path.parent.mkdir(parents=True, exist_ok=True)
        init_database(db_path)  # Pass Path object
        db = DashboardDB(db_path)  # Pass Path object
    return db


# ============================================================
# Pydantic models for API
# ============================================================


class AnalysisCreate(BaseModel):
    """Request model for creating an analysis."""

    query: str
    sources: list[str] | None = None
    limit: int = 20
    with_map: bool = True


class AlertCreate(BaseModel):
    """Request model for creating an alert/watch."""

    keywords: list[str]
    sources: list[str] = ["bodacc", "boamp", "gdelt"]


class ExportRequest(BaseModel):
    """Request model for export."""

    analysis_id: int
    format: str = "pdf"  # pdf, excel, json


# ============================================================
# Health & Status endpoints
# ============================================================


@app.get("/")
async def root():
    """Root endpoint - API info."""
    return {
        "name": "Tawiza Dashboard API",
        "version": "1.0.0",
        "endpoints": {
            "stats": "/api/stats",
            "analyses": "/api/analyses",
            "alerts": "/api/alerts",
            "export": "/api/export",
        },
    }


@app.get("/health")
async def health():
    """Health check endpoint."""
    return {"status": "ok", "timestamp": datetime.now().isoformat()}


@app.get("/api/status")
async def get_status():
    """Get overall system status."""
    db = get_db()
    stats = await get_async_stats(db)

    return {
        "status": "online",
        "mcp_server": "http://localhost:8765/sse",
        "ollama": os.getenv("OLLAMA_BASE_URL", "http://localhost:11434"),
        "stats": stats.to_dict() if stats else {},
        "timestamp": datetime.now().isoformat(),
    }


# ============================================================
# Stats endpoints
# ============================================================


@app.get("/api/stats")
async def get_stats(period: str = "last_7_days"):
    """Get dashboard statistics."""
    db = get_db()
    stats = await get_async_stats(db, period=period)
    return stats.to_dict() if stats else {}


@app.get("/api/stats/kpis")
async def get_kpis():
    """Get key performance indicators for dashboard home."""
    db = get_db()

    analyses_raw = db.get_recent_analyses(limit=1000)
    analyses = [Analysis.from_dict(a) for a in analyses_raw]
    alerts_raw = db.get_unread_alerts(limit=1000)

    total_companies = sum(a.results_count for a in analyses)
    avg_confidence = sum(a.confidence or 0 for a in analyses) / len(analyses) if analyses else 0

    return {
        "total_analyses": len(analyses),
        "total_companies": total_companies,
        "unread_alerts": len(alerts_raw),
        "avg_confidence": round(avg_confidence, 1),
    }


# ============================================================
# Analyses endpoints
# ============================================================


@app.get("/api/analyses")
async def list_analyses(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    query_filter: str | None = None,
):
    """List recent analyses."""
    db = get_db()
    analyses_raw = db.get_recent_analyses(limit=limit + offset)
    analyses = [Analysis.from_dict(a) for a in analyses_raw]

    # Apply filter if provided
    if query_filter:
        analyses = [a for a in analyses if query_filter.lower() in a.query.lower()]

    # Apply pagination
    analyses = analyses[offset : offset + limit]

    return {
        "analyses": [a.to_dict() for a in analyses],
        "total": len(analyses),
    }


@app.get("/api/analyses/{analysis_id}")
async def get_analysis(analysis_id: int):
    """Get a specific analysis by ID."""
    db = get_db()
    analysis = db.get_analysis(analysis_id)

    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    return analysis.to_dict()


@app.post("/api/analyses")
async def create_analysis(request: AnalysisCreate):
    """Create a new analysis (triggers MCP tawiza_analyze)."""
    from src.application.orchestration.data_orchestrator import DataOrchestrator
    from src.domain.debate.debate_system import DebateSystem
    from src.infrastructure.llm import create_debate_system_with_llm

    start_time = datetime.now()

    # Run orchestrator
    orchestrator = DataOrchestrator()
    orch_result = await orchestrator.search(
        query=request.query,
        limit_per_source=request.limit,
        sources=request.sources,
    )

    # Run debate
    try:
        debate = create_debate_system_with_llm(text_model="qwen3.5:27b")
    except Exception:
        debate = DebateSystem()

    all_results = [item for sr in orch_result.source_results for item in sr.results]
    debate_result = await debate.validate(
        query=request.query,
        data={"results": all_results, "sources": [sr.source for sr in orch_result.source_results]},
    )

    duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)

    # Save to database
    db = get_db()
    analysis = Analysis(
        query=request.query,
        timestamp=datetime.now(),
        sources_used=[sr.source for sr in orch_result.source_results],
        results_count=len(all_results),
        confidence=debate_result.final_confidence,
        duration_ms=duration_ms,
        metadata={
            "verdict": debate_result.verdict,
            "issues": debate_result.issues,
            "results": all_results[:50],  # Store top 50
        },
    )
    analysis_id = db.save_analysis(analysis)
    analysis.id = analysis_id

    return {
        "success": True,
        "analysis": analysis.to_dict(),
        "results_count": len(all_results),
        "confidence": debate_result.final_confidence,
    }


# ============================================================
# Alerts endpoints
# ============================================================


@app.get("/api/alerts")
async def list_alerts(
    unread_only: bool = False,
    limit: int = Query(50, ge=1, le=200),
):
    """List alerts."""
    db = get_db()
    alerts = db.get_alerts(unread_only=unread_only, limit=limit)

    return {
        "alerts": [a.to_dict() for a in alerts],
        "total": len(alerts),
        "unread": len([a for a in alerts if not a.read]),
    }


@app.post("/api/alerts/{alert_id}/read")
async def mark_alert_read(alert_id: int):
    """Mark an alert as read."""
    db = get_db()
    db.mark_alert_read(alert_id)
    return {"success": True}


@app.post("/api/alerts/read-all")
async def mark_all_alerts_read():
    """Mark all alerts as read."""
    db = get_db()
    db.mark_all_alerts_read()
    return {"success": True}


# ============================================================
# Watch/Veille endpoints
# ============================================================


@app.get("/api/watch")
async def list_watch_items():
    """List watch items (veille)."""
    db = get_db()
    items = db.get_watch_items()

    return {
        "items": [w.to_dict() for w in items],
        "total": len(items),
        "active": len([w for w in items if w.active]),
    }


@app.post("/api/watch")
async def create_watch_item(request: AlertCreate):
    """Create a new watch item."""
    db = get_db()

    watch = WatchItem(
        keywords=request.keywords,
        sources=request.sources,
        active=True,
        created_at=datetime.now(),
    )
    watch_id = db.save_watch_item(watch)
    watch.id = watch_id

    return {"success": True, "watch": watch.to_dict()}


@app.delete("/api/watch/{watch_id}")
async def delete_watch_item(watch_id: int):
    """Delete a watch item."""
    db = get_db()
    db.delete_watch_item(watch_id)
    return {"success": True}


@app.post("/api/watch/{watch_id}/toggle")
async def toggle_watch_item(watch_id: int):
    """Toggle watch item active status."""
    db = get_db()
    db.toggle_watch_item(watch_id)
    return {"success": True}


# ============================================================
# Export endpoints
# ============================================================


@app.post("/api/export")
async def export_analysis(request: ExportRequest):
    """Export an analysis to PDF/Excel/JSON."""
    db = get_db()
    analysis = db.get_analysis(request.analysis_id)

    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    export_dir = Path.home() / ".tawiza" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_query = "".join(c for c in analysis.query[:30] if c.isalnum() or c in " _-").strip()

    if request.format == "json":
        filename = f"{timestamp}_{safe_query}.json"
        filepath = export_dir / filename

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(analysis.to_dict(), f, ensure_ascii=False, indent=2, default=str)

        return {"success": True, "path": str(filepath), "filename": filename}

    elif request.format == "excel":
        try:
            import openpyxl
            from openpyxl import Workbook

            filename = f"{timestamp}_{safe_query}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"

            # Header
            ws["A1"] = "Query"
            ws["B1"] = analysis.query
            ws["A2"] = "Confidence"
            ws["B2"] = f"{analysis.confidence}%"
            ws["A3"] = "Results"
            ws["B3"] = analysis.results_count
            ws["A4"] = "Date"
            ws["B4"] = analysis.timestamp.isoformat() if analysis.timestamp else ""

            # Results
            if analysis.metadata and "results" in analysis.metadata:
                ws["A6"] = "Results"
                headers = ["Name", "SIRET", "Source", "City"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=7, column=col, value=h)

                for row, item in enumerate(analysis.metadata["results"], 8):
                    ws.cell(row=row, column=1, value=item.get("nom") or item.get("name", ""))
                    ws.cell(row=row, column=2, value=item.get("siret", ""))
                    ws.cell(row=row, column=3, value=item.get("source", ""))
                    ws.cell(row=row, column=4, value=item.get("commune", ""))

            wb.save(filepath)
            return {"success": True, "path": str(filepath), "filename": filename}

        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")

    elif request.format == "pdf":
        # Simple HTML to PDF using weasyprint or fallback to HTML
        try:
            from weasyprint import HTML

            filename = f"{timestamp}_{safe_query}.pdf"
            filepath = export_dir / filename

            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>Tawiza Analysis Report</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    h1 {{ color: #2563eb; }}
                    .kpi {{ display: inline-block; margin: 10px; padding: 15px;
                           background: #f3f4f6; border-radius: 8px; }}
                    table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background: #2563eb; color: white; }}
                </style>
            </head>
            <body>
                <h1>Tawiza Analysis Report</h1>
                <h2>{analysis.query}</h2>

                <div class="kpi">
                    <strong>Confidence</strong><br>
                    {analysis.confidence}%
                </div>
                <div class="kpi">
                    <strong>Results</strong><br>
                    {analysis.results_count}
                </div>
                <div class="kpi">
                    <strong>Sources</strong><br>
                    {len(analysis.sources_used)}
                </div>

                <h3>Sources Used</h3>
                <p>{", ".join(analysis.sources_used)}</p>

                <h3>Top Results</h3>
                <table>
                    <tr><th>Name</th><th>SIRET</th><th>Source</th><th>City</th></tr>
            """

            if analysis.metadata and "results" in analysis.metadata:
                for item in analysis.metadata["results"][:20]:
                    name = item.get("nom") or item.get("name", "N/A")
                    html_content += f"""
                    <tr>
                        <td>{name}</td>
                        <td>{item.get("siret", "")}</td>
                        <td>{item.get("source", "")}</td>
                        <td>{item.get("commune", "")}</td>
                    </tr>
                    """

            html_content += """
                </table>
                <p style="margin-top: 40px; color: #666;">
                    Generated by Tawiza Dashboard
                </p>
            </body>
            </html>
            """

            HTML(string=html_content).write_pdf(filepath)
            return {"success": True, "path": str(filepath), "filename": filename}

        except ImportError:
            # Fallback to HTML
            filename = f"{timestamp}_{safe_query}.html"
            filepath = export_dir / filename

            with open(filepath, "w", encoding="utf-8") as f:
                f.write(html_content)

            return {
                "success": True,
                "path": str(filepath),
                "filename": filename,
                "note": "PDF unavailable, exported as HTML",
            }

    else:
        raise HTTPException(status_code=400, detail=f"Unknown format: {request.format}")


@app.get("/api/export/download/{filename}")
async def download_export(filename: str):
    """Download an exported file."""
    export_dir = Path.home() / ".tawiza" / "exports"
    filepath = export_dir / filename

    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(filepath, filename=filename)


# ============================================================
# Map endpoint
# ============================================================


@app.get("/api/map/{analysis_id}")
async def get_analysis_map(analysis_id: int):
    """Get map HTML for an analysis."""
    db = get_db()
    analysis = db.get_analysis(analysis_id)

    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    if not analysis.metadata or "results" not in analysis.metadata:
        raise HTTPException(status_code=404, detail="No results for map")

    # Generate map
    import tempfile

    from src.cli.v2.agents.tools import register_all_tools
    from src.cli.v2.agents.unified.tools import ToolRegistry

    registry = ToolRegistry()
    register_all_tools(registry)

    locations = []
    for item in analysis.metadata["results"]:
        geo = item.get("geo")
        if geo and geo.get("lat"):
            locations.append(
                {
                    "nom": item.get("nom") or item.get("name", "N/A"),
                    "lat": geo["lat"],
                    "lon": geo["lon"],
                    "type": "entreprise",
                    "source": item.get("source", ""),
                }
            )

    if not locations:
        return HTMLResponse("<p>No locations to display</p>")

    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
        map_path = f.name

    await registry.execute(
        "geo.map",
        {
            "locations": locations,
            "title": f"Analysis: {analysis.query[:30]}",
            "output_path": map_path,
        },
    )

    with open(map_path) as f:
        map_html = f.read()

    return HTMLResponse(map_html)


# ============================================================
# Run server
# ============================================================


def run_api_server(host: str = "0.0.0.0", port: int = 3001):
    """Run the API server."""
    import uvicorn

    print(f"Starting Tawiza Dashboard API on http://{host}:{port}")
    uvicorn.run(app, host=host, port=port)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=3001)
    args = parser.parse_args()

    run_api_server(args.host, args.port)
