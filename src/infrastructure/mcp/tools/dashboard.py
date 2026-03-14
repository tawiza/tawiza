"""Dashboard MCP tools for Cherry Studio.

Provides tools for managing the dashboard:
- Alerts management (mark read, get details)
- Watchlist management (add/remove keywords)
- Watcher control (force poll)
"""

import json

from mcp.server.fastmcp import Context, FastMCP

from ...dashboard import DashboardDB
from ...watcher import BoampPoller, BodaccPoller, GdeltPoller, WatcherStorage

# Global instances
_db: DashboardDB | None = None
_storage: WatcherStorage | None = None


def get_db() -> DashboardDB:
    """Get or create dashboard database instance."""
    global _db
    if _db is None:
        _db = DashboardDB()
    return _db


def get_storage() -> WatcherStorage:
    """Get or create watcher storage instance."""
    global _storage
    if _storage is None:
        _storage = WatcherStorage(get_db())
    return _storage


def register_dashboard_tools(mcp: FastMCP) -> None:
    """Register dashboard tools on the MCP server."""

    # =========== Alerts Management ===========

    @mcp.tool()
    async def dashboard_mark_read(
        alert_ids: list[int] | None = None,
        mark_all: bool = False,
        ctx: Context = None,
    ) -> str:
        """Marque des alertes comme lues.

        Args:
            alert_ids: Liste des IDs d'alertes à marquer (optionnel)
            mark_all: Si True, marque toutes les alertes comme lues

        Returns:
            Nombre d'alertes marquées
        """
        db = get_db()

        if ctx:
            ctx.info("[Dashboard] Marking alerts as read...")

        count = await db.async_mark_alerts_read(alert_ids=alert_ids, all=mark_all)

        if ctx:
            ctx.info(f"[Dashboard] Marked {count} alerts as read")

        return json.dumps(
            {
                "success": True,
                "marked_count": count,
                "message": f"{count} alertes marquées comme lues",
            },
            ensure_ascii=False,
        )

    @mcp.tool()
    async def dashboard_get_alert(
        alert_id: int,
        ctx: Context = None,
    ) -> str:
        """Récupère le détail complet d'une alerte.

        Args:
            alert_id: ID de l'alerte

        Returns:
            Détails complets de l'alerte
        """
        db = get_db()

        if ctx:
            ctx.info(f"[Dashboard] Getting alert {alert_id}...")

        alert = db.get_alert_detail(alert_id)

        if not alert:
            return json.dumps(
                {
                    "success": False,
                    "error": f"Alerte {alert_id} non trouvée",
                },
                ensure_ascii=False,
            )

        return json.dumps(
            {
                "success": True,
                "alert": alert,
            },
            ensure_ascii=False,
            indent=2,
        )

    @mcp.tool()
    async def dashboard_get_alerts(
        source: str | None = None,
        limit: int = 20,
        include_read: bool = False,
        ctx: Context = None,
    ) -> str:
        """Récupère les alertes du dashboard.

        Args:
            source: Filtrer par source (bodacc, boamp, gdelt)
            limit: Nombre maximum d'alertes
            include_read: Inclure les alertes déjà lues

        Returns:
            Liste des alertes
        """
        db = get_db()

        if ctx:
            ctx.info("[Dashboard] Getting alerts...")

        if include_read:
            # Get all alerts
            cursor = db.conn.execute(
                """
                SELECT id, source, type, title, content, url, detected_at, read
                FROM alerts
                WHERE (? IS NULL OR source = ?)
                ORDER BY detected_at DESC
                LIMIT ?
                """,
                (source, source, limit),
            )
        else:
            alerts = db.get_unread_alerts(limit=limit)
            if source:
                alerts = [a for a in alerts if a["source"] == source]
            return json.dumps(
                {
                    "success": True,
                    "count": len(alerts),
                    "alerts": alerts,
                },
                ensure_ascii=False,
                indent=2,
            )

        alerts = [dict(row) for row in cursor.fetchall()]

        return json.dumps(
            {
                "success": True,
                "count": len(alerts),
                "alerts": alerts,
            },
            ensure_ascii=False,
            indent=2,
        )

    # =========== Watchlist Management ===========

    @mcp.tool()
    async def watchlist_add(
        keywords: list[str],
        sources: list[str] | None = None,
        ctx: Context = None,
    ) -> str:
        """Ajoute des mots-clés à surveiller.

        Args:
            keywords: Liste de mots-clés à surveiller
            sources: Sources à surveiller (bodacc, boamp, gdelt). Défaut: toutes

        Returns:
            Confirmation d'ajout
        """
        db = get_db()

        if not keywords:
            return json.dumps(
                {
                    "success": False,
                    "error": "Aucun mot-clé fourni",
                },
                ensure_ascii=False,
            )

        if sources is None:
            sources = ["bodacc", "boamp", "gdelt"]

        if ctx:
            ctx.info(f"[Dashboard] Adding keywords: {keywords}")

        item_id = await db.async_add_watchlist_item(keywords, sources)

        if ctx:
            ctx.info(f"[Dashboard] Watchlist item created: {item_id}")

        return json.dumps(
            {
                "success": True,
                "id": item_id,
                "keywords": keywords,
                "sources": sources,
                "message": f"Mots-clés ajoutés à la watchlist: {', '.join(keywords)}",
            },
            ensure_ascii=False,
        )

    @mcp.tool()
    async def watchlist_remove(
        keywords: list[str],
        ctx: Context = None,
    ) -> str:
        """Retire des mots-clés de la surveillance.

        Args:
            keywords: Liste de mots-clés à retirer

        Returns:
            Confirmation de suppression
        """
        db = get_db()

        if not keywords:
            return json.dumps(
                {
                    "success": False,
                    "error": "Aucun mot-clé fourni",
                },
                ensure_ascii=False,
            )

        if ctx:
            ctx.info(f"[Dashboard] Removing keywords: {keywords}")

        count = db.remove_watchlist_keywords(keywords)

        return json.dumps(
            {
                "success": True,
                "removed_count": count,
                "keywords": keywords,
                "message": f"{count} entrées de watchlist désactivées",
            },
            ensure_ascii=False,
        )

    @mcp.tool()
    async def watchlist_list(ctx: Context = None) -> str:
        """Liste tous les mots-clés surveillés.

        Returns:
            Liste des entrées de la watchlist
        """
        db = get_db()

        if ctx:
            ctx.info("[Dashboard] Getting watchlist...")

        watchlist = await db.async_get_watchlist(active_only=True)

        # Flatten keywords
        all_keywords = set()
        for item in watchlist:
            all_keywords.update(item["keywords"])

        return json.dumps(
            {
                "success": True,
                "items": watchlist,
                "all_keywords": sorted(all_keywords),
                "total_items": len(watchlist),
            },
            ensure_ascii=False,
            indent=2,
        )

    # =========== Watcher Control ===========

    @mcp.tool()
    async def watcher_force_poll(
        source: str | None = None,
        ctx: Context = None,
    ) -> str:
        """Force un poll immédiat des sources.

        Args:
            source: Source spécifique (bodacc, boamp, gdelt) ou None pour toutes

        Returns:
            Résultat du polling avec nombre d'alertes trouvées
        """
        storage = get_storage()

        sources_to_poll = [source] if source else ["bodacc", "boamp", "gdelt"]

        if ctx:
            ctx.info(f"[Dashboard] Force polling: {sources_to_poll}")
            ctx.report_progress(0, len(sources_to_poll), "Starting poll...")

        results = {}
        pollers = {
            "bodacc": BodaccPoller(),
            "boamp": BoampPoller(),
            "gdelt": GdeltPoller(),
        }

        for i, src in enumerate(sources_to_poll):
            if src not in pollers:
                results[src] = {"error": f"Unknown source: {src}"}
                continue

            if ctx:
                ctx.info(f"[Dashboard] Polling {src}...")
                ctx.report_progress(i, len(sources_to_poll), f"Polling {src}...")

            # Get keywords for this source
            keywords = await storage.async_get_keywords_for_source(src)

            # Poll
            poller = pollers[src]
            alerts, error = await poller.safe_poll(keywords)

            if error:
                results[src] = {"error": error, "alerts_found": 0}
                continue

            # Save new alerts
            saved = 0
            for alert in alerts:
                exists = await storage.async_alert_exists(src, alert.title, alert.url)
                if not exists:
                    await storage.async_save_alert(alert)
                    saved += 1

            results[src] = {
                "alerts_found": len(alerts),
                "alerts_saved": saved,
                "keywords_used": keywords,
            }

            # Update poll status
            await storage.async_record_poll(src, poller.default_interval)

        if ctx:
            ctx.report_progress(len(sources_to_poll), len(sources_to_poll), "Poll complete")

        total_found = sum(r.get("alerts_found", 0) for r in results.values() if "error" not in r)
        total_saved = sum(r.get("alerts_saved", 0) for r in results.values() if "error" not in r)

        return json.dumps(
            {
                "success": True,
                "results": results,
                "summary": {
                    "total_alerts_found": total_found,
                    "total_alerts_saved": total_saved,
                },
            },
            ensure_ascii=False,
            indent=2,
        )

    # =========== Dashboard Status & KPIs ===========

    @mcp.tool()
    async def tawiza_dashboard_status(ctx: Context = None) -> str:
        """Récupère le status global du dashboard Tawiza.

        Retourne les KPIs principaux:
        - Nombre d'analyses effectuées
        - Nombre d'entreprises trouvées
        - Alertes non lues
        - Confiance moyenne

        Returns:
            KPIs du dashboard au format JSON
        """
        db = get_db()

        if ctx:
            ctx.info("[Dashboard] Getting status and KPIs...")

        # Get analyses stats
        analyses = db.get_recent_analyses(limit=1000)
        total_analyses = len(analyses)
        total_companies = sum(a.get("results_count", 0) for a in analyses)
        avg_confidence = (
            sum(a.get("confidence", 0) or 0 for a in analyses) / total_analyses
            if total_analyses
            else 0
        )

        # Get alerts
        alerts_count = db.get_alerts_count()

        # Get recent analyses for display
        recent = db.get_recent_analyses(limit=5)

        return json.dumps(
            {
                "success": True,
                "kpis": {
                    "total_analyses": total_analyses,
                    "total_companies": total_companies,
                    "avg_confidence": round(avg_confidence, 1),
                    "unread_alerts": alerts_count.get("total_unread", 0),
                },
                "recent_analyses": recent,
                "alerts_by_source": alerts_count.get("by_source", {}),
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        )

    @mcp.tool()
    async def tawiza_history(
        limit: int = 20,
        query_filter: str | None = None,
        ctx: Context = None,
    ) -> str:
        """Récupère l'historique des analyses.

        Args:
            limit: Nombre maximum d'analyses à retourner
            query_filter: Filtrer par texte dans la requête

        Returns:
            Liste des analyses passées avec métadonnées
        """
        db = get_db()

        if ctx:
            ctx.info(f"[Dashboard] Getting history (limit={limit})...")

        analyses = db.get_recent_analyses(limit=limit)

        # Apply filter if provided
        if query_filter:
            analyses = [a for a in analyses if query_filter.lower() in a.get("query", "").lower()]

        return json.dumps(
            {
                "success": True,
                "count": len(analyses),
                "analyses": analyses,
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        )

    @mcp.tool()
    async def tawiza_export(
        analysis_id: int,
        format: str = "json",
        ctx: Context = None,
    ) -> str:
        """Exporte une analyse en PDF, Excel ou JSON.

        Args:
            analysis_id: ID de l'analyse à exporter
            format: Format d'export (json, excel, pdf)

        Returns:
            Chemin du fichier exporté ou données JSON
        """
        from datetime import datetime
        from pathlib import Path

        db = get_db()

        if ctx:
            ctx.info(f"[Dashboard] Exporting analysis {analysis_id} as {format}...")

        # Get full analysis data
        cursor = db.conn.execute(
            "SELECT * FROM analyses WHERE id = ?",
            (analysis_id,),
        )
        row = cursor.fetchone()

        if not row:
            return json.dumps(
                {
                    "success": False,
                    "error": f"Analyse {analysis_id} non trouvée",
                },
                ensure_ascii=False,
            )

        analysis = {
            "id": row["id"],
            "query": row["query"],
            "timestamp": row["timestamp"],
            "sources_used": json.loads(row["sources_used"]) if row["sources_used"] else [],
            "results_count": row["results_count"],
            "confidence": row["confidence"],
            "duration_ms": row["duration_ms"],
            "metadata": json.loads(row["metadata"]) if row["metadata"] else {},
        }

        # Export directory
        export_dir = Path.home() / ".tawiza" / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_query = (
            "".join(c for c in analysis["query"][:30] if c.isalnum() or c in " _-")
            .strip()
            .replace(" ", "_")
        )

        if format == "json":
            filename = f"{timestamp}_{safe_query}.json"
            filepath = export_dir / filename
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(analysis, f, ensure_ascii=False, indent=2, default=str)

            return json.dumps(
                {
                    "success": True,
                    "format": "json",
                    "path": str(filepath),
                    "filename": filename,
                },
                ensure_ascii=False,
            )

        elif format == "excel":
            try:
                from openpyxl import Workbook

                filename = f"{timestamp}_{safe_query}.xlsx"
                filepath = export_dir / filename

                wb = Workbook()
                ws = wb.active
                ws.title = "Analysis"

                # Header info
                ws["A1"], ws["B1"] = "Query", analysis["query"]
                ws["A2"], ws["B2"] = "Confidence", f"{analysis['confidence']}%"
                ws["A3"], ws["B3"] = "Results", analysis["results_count"]
                ws["A4"], ws["B4"] = "Date", analysis["timestamp"]
                ws["A5"], ws["B5"] = "Sources", ", ".join(analysis["sources_used"])

                # Results table
                results = analysis.get("metadata", {}).get("results", [])
                if results:
                    ws["A7"] = "Entreprises"
                    headers = ["Nom", "SIRET", "Source", "Ville"]
                    for col, h in enumerate(headers, 1):
                        ws.cell(row=8, column=col, value=h)

                    for row_idx, item in enumerate(results[:100], 9):
                        ws.cell(
                            row=row_idx, column=1, value=item.get("nom") or item.get("name", "")
                        )
                        ws.cell(row=row_idx, column=2, value=item.get("siret", ""))
                        ws.cell(row=row_idx, column=3, value=item.get("source", ""))
                        ws.cell(row=row_idx, column=4, value=item.get("commune", ""))

                wb.save(filepath)

                return json.dumps(
                    {
                        "success": True,
                        "format": "excel",
                        "path": str(filepath),
                        "filename": filename,
                        "rows": len(results),
                    },
                    ensure_ascii=False,
                )

            except ImportError:
                return json.dumps(
                    {
                        "success": False,
                        "error": "openpyxl non installé - utilisez: pip install openpyxl",
                    },
                    ensure_ascii=False,
                )

        elif format == "pdf":
            try:
                from weasyprint import HTML

                filename = f"{timestamp}_{safe_query}.pdf"
                filepath = export_dir / filename

                results = analysis.get("metadata", {}).get("results", [])

                html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    body {{ font-family: Arial, sans-serif; margin: 40px; }}
    h1 {{ color: #2563eb; }}
    .kpi {{ display: inline-block; margin: 10px; padding: 15px; background: #f3f4f6; border-radius: 8px; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
    th {{ background: #2563eb; color: white; }}
</style></head><body>
<h1>Tawiza - Rapport d'Analyse</h1>
<h2>{analysis["query"]}</h2>
<div class="kpi"><strong>Confiance</strong><br>{analysis["confidence"]}%</div>
<div class="kpi"><strong>Résultats</strong><br>{analysis["results_count"]}</div>
<div class="kpi"><strong>Sources</strong><br>{len(analysis["sources_used"])}</div>
<h3>Sources utilisées</h3>
<p>{", ".join(analysis["sources_used"])}</p>
<h3>Entreprises trouvées</h3>
<table><tr><th>Nom</th><th>SIRET</th><th>Source</th><th>Ville</th></tr>"""

                for item in results[:50]:
                    name = item.get("nom") or item.get("name", "N/A")
                    html += f"<tr><td>{name}</td><td>{item.get('siret', '')}</td><td>{item.get('source', '')}</td><td>{item.get('commune', '')}</td></tr>"

                html += """</table>
<p style="margin-top: 40px; color: #666;">Généré par Tawiza Dashboard</p>
</body></html>"""

                HTML(string=html).write_pdf(filepath)

                return json.dumps(
                    {
                        "success": True,
                        "format": "pdf",
                        "path": str(filepath),
                        "filename": filename,
                    },
                    ensure_ascii=False,
                )

            except ImportError:
                return json.dumps(
                    {
                        "success": False,
                        "error": "weasyprint non installé - utilisez: pip install weasyprint",
                    },
                    ensure_ascii=False,
                )

        else:
            return json.dumps(
                {
                    "success": False,
                    "error": f"Format inconnu: {format}. Utilisez: json, excel, pdf",
                },
                ensure_ascii=False,
            )

    @mcp.tool()
    async def watcher_status(ctx: Context = None) -> str:
        """Récupère le status du watcher.

        Returns:
            Status de chaque source avec prochains polls prévus
        """
        db = get_db()
        storage = get_storage()

        if ctx:
            ctx.info("[Dashboard] Getting watcher status...")

        poll_status = db.get_poll_status()
        next_polls = storage.get_next_poll_times()

        sources = {}
        for src in ["bodacc", "boamp", "gdelt"]:
            ps = poll_status.get(src, {})
            sources[src] = {
                "last_poll": ps.get("last_poll"),
                "next_poll": next_polls.get(src, "unknown"),
                "polls_count": ps.get("polls_count", 0),
                "last_error": ps.get("last_error"),
            }

        return json.dumps(
            {
                "success": True,
                "sources": sources,
            },
            ensure_ascii=False,
            indent=2,
        )
